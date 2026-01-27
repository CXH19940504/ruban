import openpyxl
import logging
import os
from datetime import datetime
from ruban.utils.log import get_logger


logger = get_logger(__name__)


class ExcelHandler:
    def __init__(self, excel_path):
        self.excel_path = excel_path

    def read_excel_data(self, sheet_name, headers):
        logger.info(f"开始读取Excel文件: {self.excel_path}, 工作表: {sheet_name}, 表头: {headers}")

        try:
            # 加载Excel文件
            workbook = openpyxl.load_workbook(self.excel_path)
            logger.info(f"成功加载Excel文件: {self.excel_path}")

            # 检查工作表是否存在
            if sheet_name not in workbook.sheetnames:
                error_msg = f"工作表 '{sheet_name}' 不存在"
                logger.error(error_msg)
                raise Exception(error_msg)

            # 选择工作表
            sheet = workbook[sheet_name]
            logger.info(f"成功选择工作表: {sheet_name}")

            column_no = {}
            data = {}
            for header_name in headers:
                # 查找表头所在列
                header_column = None
                for cell in sheet[1]:  # 假设表头在第一行
                    if cell.value == header_name:
                        header_column = cell.column
                        logger.info(f"找到表头 '{header_name}' 在第 {header_column} 列")
                        break

                if not header_column:
                    error_msg = f"表头 '{header_name}' 未找到"
                    logger.error(error_msg)
                    raise Exception(error_msg)
                column_no[header_name] = header_column
                data[header_name] = []

            # 从表头下方读取所有数据
            for row in sheet.iter_rows(min_row=2):
                for header_name in headers:
                    _value = row[column_no[header_name]-1].value
                    if _value:
                        data[header_name].append(str(_value).strip())

            logger.info(f"成功读取 {len(data[headers[0]])} 条数据")
            return data

        except Exception as e:
            error_msg = f"读取Excel文件时发生错误: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise Exception(error_msg)

    def write_excel_data(self, sheet_name, data):
        count = len(data) if data else 0
        logger.info(f"开始写入Excel文件: {self.excel_path}, 工作表: {sheet_name}, 表头: value, 数据量: {count}")

        try:
            if not data:
                error_msg = '数据列表不能为空'
                logger.error(error_msg)
                return {'status': 'error', 'message': error_msg}

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
            logger.info(f"成功创建工作表: {sheet_name}")

            sheet.cell(row=1, column=1, value="value")
            # 写入数据（不清空原有数据，直接覆盖/追加）
            for idx, value in enumerate(data, start=2):
                sheet.cell(row=idx, column=1, value=value)

            workbook.save(self.excel_path)
            logger.info(f"成功保存Excel文件: {self.excel_path}")

            success_msg = f"成功写入 {len(data)} 条数据到Excel文件"
            logger.info(success_msg)
            return {'status': 'success', 'message': success_msg}

        except Exception as e:
            error_msg = f"写入Excel文件时发生错误: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {'status': 'error', 'message': error_msg}

    def write_multiple_columns(self, sheet_name, columns_data: dict, headers=None):
        """
        写入多列数据到Excel文件

        参数:
            self.excel_path: Excel文件路径
            sheet_name: 工作表名称
            headers: 工作表的表头
            columns_data: 字典，键为列名，值为数据列表
        """
        count = len(columns_data) if columns_data else 0
        logger.info(f"开始写入多列数据到Excel文件: {self.excel_path}, 工作表: {sheet_name}, 列数: {count}")

        try:
            if not columns_data:
                error_msg = '列数据字典不能为空'
                logger.error(error_msg)
                return {'status': 'error', 'message': error_msg}

            if not headers:
                headers = columns_data.keys()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
            logger.info(f"成功创建工作表: {sheet_name}")

            # 处理每一行数据
            for header_column, header_name in enumerate(headers, start=1):
                data = columns_data[header_name]
                sheet.cell(row=1, column=header_column, value=header_name)
                logger.info(f"创建表头'{header_name}', 数据量: {len(data) if data else 0}")

                # 写入数据（从第2行开始）
                for idx, value in enumerate(data, start=2):
                    sheet.cell(row=idx, column=header_column, value=value)

            workbook.save(self.excel_path)
            logger.info(f"成功保存Excel文件: {self.excel_path}")

            success_msg = f"成功写入 {len(columns_data)} 列数据到Excel文件"
            logger.info(success_msg)
            return {'status': 'success', 'message': success_msg}

        except Exception as e:
            error_msg = f"写入多列数据到Excel文件时发生错误: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {'status': 'error', 'message': error_msg}
